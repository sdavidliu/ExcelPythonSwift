import openpyxl
import re

def zonePrimaryDirections(sheet) -> str:
    answer = ""
    primarySet = set()
    zpDirection = sheet
    startZone = ""
    name = ""
    for i in range(2,zpDirection.max_row+1):
        if (zpDirection.cell(row=i, column=1).value != None):
            startZone = zpDirection.cell(row=i, column=1).value
            name = 'z' + startZone[5:]
            if ('Out' in startZone):
                name = 'oz' + startZone[8:]
            primary = zpDirection.cell(row=i, column=2).value
            d1 = getStr(zpDirection.cell(row=i, column=3).value)
            d1arrow = zpDirection.cell(row=i, column=4).value
            d1image = zpDirection.cell(row=i, column=5).value
            if d1image == None:
                d1image = ""
            d2 = getStr(zpDirection.cell(row=i, column=6).value)
            d2arrow = zpDirection.cell(row=i, column=7).value
            d2image = zpDirection.cell(row=i, column=8).value
            if d2image == None:
                d2image = ""
            d3 = getStr(zpDirection.cell(row=i, column=9).value)
            d3arrow = zpDirection.cell(row=i, column=10).value
            d3image = zpDirection.cell(row=i, column=11).value
            if d3image == None:
                d3image = ""

            if not (startZone in primarySet):
                answer += 'var ' + name + 'Neighbors = [String:[PathStep]]()\n'
                primarySet.add(startZone)
            if (d1 != 'N/A'):
                answer += 'tempPathStepList = [PathStep]()\ntempPathStepList.append(PathStep(directionText: ' + d1 + ', directionImage: "' + d1image + '", arrow: .' + convertArrow(d1arrow) + '))\n'
            if (d2 != 'N/A'):
                answer += 'tempPathStepList.append(PathStep(directionText: ' + d2 + ', directionImage: "' + d2image + '", arrow: .' + convertArrow(d2arrow) + '))\n'
            if (d3 != 'N/A'):
                answer += 'tempPathStepList.append(PathStep(directionText: ' + d3 + ', directionImage: "' + d3image + '", arrow: .' + convertArrow(d3arrow) + '))\n'
            answer += name + 'Neighbors["' + primary + '"] = tempPathStepList\n\n'

    return answer


def zoneSecondaryDirections(sheet) -> str:
    answer = ""
    zsDirection = sheet
    startZone = ""
    name = ""
    for i in range(2,zsDirection.max_row+1):
        if (zsDirection.cell(row=i, column=1).value != None):
            startZone = zsDirection.cell(row=i, column=1).value
            name = 'z' + startZone[5:]
            if ('Out' in startZone):
                name = 'oz' + startZone[8:]
            primary = zsDirection.cell(row=i, column=2).value
            d1 = getStr(zsDirection.cell(row=i, column=3).value)
            d1arrow = zsDirection.cell(row=i, column=4).value
            d1image = zsDirection.cell(row=i, column=5).value
            if d1image == None:
                d1image = ""
            d2 = getStr(zsDirection.cell(row=i, column=6).value)
            d2arrow = zsDirection.cell(row=i, column=7).value
            d2image = zsDirection.cell(row=i, column=8).value
            if d2image == None:
                d2image = ""
            d3 = getStr(zsDirection.cell(row=i, column=9).value)
            d3arrow = zsDirection.cell(row=i, column=10).value
            d3image = zsDirection.cell(row=i, column=11).value
            if d3image == None:
                d3image = ""

            if (d1 != 'N/A'):
                answer += 'tempPathStepList = [PathStep]()\ntempPathStepList.append(PathStep(directionText: ' + d1 + ', directionImage: "' + d1image + '", arrow: .' + convertArrow(d1arrow) + '))\n'
            if (d2 != 'N/A'):
                answer += 'tempPathStepList.append(PathStep(directionText: ' + d2 + ', directionImage: "' + d2image + '", arrow: .' + convertArrow(d2arrow) + '))\n'
            if (d3 != 'N/A'):
                answer += 'tempPathStepList.append(PathStep(directionText: ' + d3 + ', directionImage: "' + d3image + '", arrow: .' + convertArrow(d3arrow) + '))\n'
            answer += name + 'Neighbors["' + primary + '"] = tempPathStepList\n\n'

    return answer


def primaryPrimaryDirections(sheet,maxDirections) -> str:
    answer = ""
    primarySet = set()
    ppDirection = sheet
    startPrimary = ""
    name = ""

    for i in range(2,ppDirection.max_row+1):
        if (ppDirection.cell(row=i, column=1).value != None):
            startPrimary = ppDirection.cell(row=i, column=1).value.lower().strip()
            endPrimary = ppDirection.cell(row=i, column=2).value
            d1 = getStr(ppDirection.cell(row=i, column=3).value.strip())
            d1arrow = ppDirection.cell(row=i, column=4).value.strip()
            d1image = ppDirection.cell(row=i, column=5).value
            if d1image == None:
                d1image = ""
            d2 = getStr(ppDirection.cell(row=i, column=6).value.strip())
            d2arrow = ppDirection.cell(row=i, column=7).value.strip()
            d2image = ppDirection.cell(row=i, column=8).value
            if d2image == None:
                d2image = ""
            d3 = getStr(ppDirection.cell(row=i, column=9).value.strip())
            d3arrow = ppDirection.cell(row=i, column=10).value.strip()
            d3image = ppDirection.cell(row=i, column=11).value
            if d3image == None:
                d3image = ""
            d4 = getStr(ppDirection.cell(row=i, column=12).value)
            d4arrow = ppDirection.cell(row=i, column=13).value
            d4image = ppDirection.cell(row=i, column=14).value
            if d4image == None:
                d4image = ""
            
            if not (startPrimary in primarySet):
                answer += 'var ' + startPrimary + 'Neighbors = [String:[PathStep]]()\n'
                primarySet.add(startPrimary)
            if (d1 != 'N/A'):
                answer += 'tempPathStepList = [PathStep]()\ntempPathStepList.append(PathStep(directionText: ' + d1 + ', directionImage: "' + d1image + '", arrow: .' + convertArrow(d1arrow) + '))\n'
            else:
                answer += 'tempPathStepList = [PathStep]()\n'
            if (d2 != 'N/A'):
                answer += 'tempPathStepList.append(PathStep(directionText: ' + d2 + ', directionImage: "' + d2image + '", arrow: .' + convertArrow(d2arrow) + '))\n'
            if (d3 != 'N/A'):
                answer += 'tempPathStepList.append(PathStep(directionText: ' + d3 + ', directionImage: "' + d3image + '", arrow: .' + convertArrow(d3arrow) + '))\n'
            if (d4 != 'N/A' and maxDirections >= 4):
                answer += 'tempPathStepList.append(PathStep(directionText: ' + d4 + ', directionImage: "' + d4image + '", arrow: .' + convertArrow(d4arrow) + '))\n'
            answer += startPrimary + 'Neighbors["' + endPrimary + '"] = tempPathStepList\n\n'

    return answer



def primarySecondaryDirections(sheet,maxDirections) -> str:
    answer = "\n"
    primarySet = set()
    psDirection = sheet
    startPrimary = ""
    name = ""

    for i in range(2,psDirection.max_row+1):
        if (psDirection.cell(row=i, column=1).value != None):
            startPrimary = psDirection.cell(row=i, column=1).value.lower().strip()
            endPrimary = psDirection.cell(row=i, column=2).value
            d1 = getStr(psDirection.cell(row=i, column=3).value)
            d1arrow = psDirection.cell(row=i, column=4).value
            d1image = psDirection.cell(row=i, column=5).value
            if d1image == None:
                d1image = ""
            d2 = getStr(psDirection.cell(row=i, column=6).value)
            d2arrow = psDirection.cell(row=i, column=7).value
            d2image = psDirection.cell(row=i, column=8).value
            if d2image == None:
                d2image = ""
            d3 = getStr(psDirection.cell(row=i, column=9).value)
            d3arrow = psDirection.cell(row=i, column=10).value
            d3image = psDirection.cell(row=i, column=11).value
            if d3image == None:
                d3image = ""
            d4 = getStr(psDirection.cell(row=i, column=12).value)
            d4arrow = psDirection.cell(row=i, column=13).value
            d4image = psDirection.cell(row=i, column=14).value
            if d4image == None:
                d4image = ""

            if (d1 != 'N/A' and d1 != None):
                answer += 'tempPathStepList = [PathStep]()\ntempPathStepList.append(PathStep(directionText: ' + d1 + ', directionImage: "' + d1image + '", arrow: .' + convertArrow(d1arrow) + '))\n'
            else:
                answer += 'tempPathStepList = [PathStep]()\n'
            if (d2 != 'N/A' and d2 != None):
                answer += 'tempPathStepList.append(PathStep(directionText: ' + d2 + ', directionImage: "' + d2image + '", arrow: .' + convertArrow(d2arrow) + '))\n'
            if (d3 != 'N/A' and d3 != None):
                answer += 'tempPathStepList.append(PathStep(directionText: ' + d3 + ', directionImage: "' + d3image + '", arrow: .' + convertArrow(d3arrow) + '))\n'
            if (d4 != 'N/A' and d4 != None and maxDirections >= 4):
                answer += 'tempPathStepList.append(PathStep(directionText: ' + d4 + ', directionImage: "' + d4image + '", arrow: .' + convertArrow(d4arrow) + '))\n'
            answer += startPrimary + 'Neighbors["' + endPrimary + '"] = tempPathStepList\n\n'

    return answer


def addToPrimaryPrimaryMap(wb) -> str:
    answer = ""

    zoneSet = set()
    f1z = wb.get_sheet_by_name('F1 Zone-Primary Start Direction')
    for i in range(2,f1z.max_row+1):
        if (f1z.cell(row=i, column=1).value != None):
            zoneSet.add(f1z.cell(row=i, column=1).value)
    f2z = wb.get_sheet_by_name('F2 Zone-Primary Start Direction')
    for i in range(2,f2z.max_row+1):
        if (f2z.cell(row=i, column=1).value != None):
            zoneSet.add(f2z.cell(row=i, column=1).value)
    f3z = wb.get_sheet_by_name('F3 Zone-Primary Start Direction')
    for i in range(2,f3z.max_row+1):
        if (f3z.cell(row=i, column=1).value != None):
            zoneSet.add(f3z.cell(row=i, column=1).value)
    for z in zoneSet:
        if ('Out' in z):
            name = 'oz' + z[8:]
            answer += 'primaryToPrimaryDescriptionsMap["' + z + '"] = ' + name.lower().replace(' ','') + 'Neighbors\n'
        else:
            answer += 'primaryToPrimaryDescriptionsMap["' + z + '"] = z' + z.lower().replace(' ','')[4:] + 'Neighbors\n'
    f1p = wb.get_sheet_by_name('F1 Primary List')
    for i in range(2,f1p.max_row+1):
        if (f1p.cell(row=i, column=1).value != None):
            answer += 'primaryToPrimaryDescriptionsMap["' + f1p.cell(row=i, column=1).value + '"] = ' + f1p.cell(row=i, column=1).value.lower() + 'Neighbors\n'
    f2p = wb.get_sheet_by_name('F2 Primary List')
    for i in range(2,f2p.max_row+1):
        if (f2p.cell(row=i, column=1).value != None):
            answer += 'primaryToPrimaryDescriptionsMap["' + f2p.cell(row=i, column=1).value + '"] = ' + f2p.cell(row=i, column=1).value.lower() + 'Neighbors\n'
    f3p = wb.get_sheet_by_name('F3 Primary List')
    for i in range(2,f3p.max_row+1):
        if (f3p.cell(row=i, column=1).value != None):
            answer += 'primaryToPrimaryDescriptionsMap["' + f3p.cell(row=i, column=1).value + '"] = ' + f3p.cell(row=i, column=1).value.lower() + 'Neighbors\n'
    return answer
    


def convertArrow(arrow) -> str:
    arrow = arrow.lower().strip()
    if (arrow.strip() == 'forward'):
        return 'forward'
    if (arrow == 'right'):
        return 'right'
    if (arrow == 'left'):
        return 'left'
    if (arrow == 'slight right'):
        return 'slightRight'
    if (arrow == 'slight left'):
        return 'slightLeft'
    if (arrow == 'right u-turn'):
        return 'rightUTurn'
    if (arrow == 'left u-turn'):
        return 'leftUTurn'
    return 'unknown'


def getStr(word):
    if word == None or word == 'N/A':
        return word
    try:
        answer = re.findall(r'"([^"]*)"',word)[0]
    except:
        return 'N/A'
    return '"' + answer + '"'


if __name__ == '__main__':
    wb = openpyxl.load_workbook('Wayfinder Master Spreadsheet.xlsx')
    #print(wb.get_sheet_names())
    f1zp = zonePrimaryDirections(wb.get_sheet_by_name('F1 Zone-Primary Start Direction'))
    f2zp = zonePrimaryDirections(wb.get_sheet_by_name('F2 Zone-Primary Start Direction'))
    f3zp = zonePrimaryDirections(wb.get_sheet_by_name('F3 Zone-Primary Start Direction'))
    f1zs = zoneSecondaryDirections(wb.get_sheet_by_name('F1 Zone-Second. Start Direction'))
    f2zs = zoneSecondaryDirections(wb.get_sheet_by_name('F2 Zone-Second. Start Direction'))
    f3zs = zoneSecondaryDirections(wb.get_sheet_by_name('F3 Zone-Second. Start Direction'))
    f1pp = primaryPrimaryDirections(wb.get_sheet_by_name('F1 Primary-Primary Directions'),4)
    f2pp = primaryPrimaryDirections(wb.get_sheet_by_name('F2 Primary-Primary Directions'),3)
    f3pp = primaryPrimaryDirections(wb.get_sheet_by_name('F3 Primary-Primary Directions'),4)
    f1ps = primarySecondaryDirections(wb.get_sheet_by_name('F1 Primary-Second. Directions'),3)
    f2ps = primarySecondaryDirections(wb.get_sheet_by_name('F2 Primary-Second. Directions'),4)
    f3ps = primarySecondaryDirections(wb.get_sheet_by_name('F3 Primary-Second. Directions'),4)
    p = addToPrimaryPrimaryMap(wb)
    # print(zpDirection.max_row)
    # print(zpDirection.max_column)
    file = open('NavigationResult.txt','w')
    file.write('var tempPathStepList = [PathStep]()\n\n')
    file.write(f1zp)
    file.write(f2zp)
    file.write(f3zp)
    file.write(f1zs)
    file.write(f2zs)
    file.write(f3zs)
    file.write(f1pp)
    file.write(f2pp)
    file.write(f3pp)
    file.write(f1ps)
    file.write(f2ps)
    file.write(f3ps)
    file.write(p)
    file.close()
