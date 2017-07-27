import openpyxl

def createPrimaryExit(wb) -> str:
    answer = '\n'
    f1Sheet = wb.get_sheet_by_name('F1 Primary List')
    for i in range(2,f1Sheet.max_row+1):
        pe = f1Sheet.cell(row=i, column=1).value
        if (pe != None):
            answer += 'let ' + pe.lower() + ' = PrimaryExit(strId: "' + pe + '")\n'
    f2Sheet = wb.get_sheet_by_name('F2 Primary List')
    for i in range(2,f2Sheet.max_row+1):
        pe = f2Sheet.cell(row=i, column=1).value
        if (pe != None):
            answer += 'let ' + pe.lower() + ' = PrimaryExit(strId: "' + pe + '")\n'
    f3Sheet = wb.get_sheet_by_name('F3 Primary List')
    for i in range(2,f3Sheet.max_row+1):
        pe = f3Sheet.cell(row=i, column=1).value
        if (pe != None):
            answer += 'let ' + pe.lower() + ' = PrimaryExit(strId: "' + pe + '")\n'
    return answer

def createZonePrimaries(wb) -> str:
    answer = '\n'
    f1Zone = wb.get_sheet_by_name('F1 Zone Info')
    for i in range(2,f1Zone.max_row + 1):
        zone = "".join(f1Zone.cell(row=i, column=1).value.split()).lower()
        primary = f1Zone.cell(row=i, column=2).value.strip(' ').lower()
        answer += 'let ' + zone + 'Primaries = [' + primary + ']\n'
    f2Zone = wb.get_sheet_by_name('F2 Zone Info')
    for i in range(2,f2Zone.max_row + 1):
        zone = "".join(f2Zone.cell(row=i, column=1).value.split()).lower()
        primary = f2Zone.cell(row=i, column=2).value.strip(' ').lower()
        answer += 'let ' + zone + 'Primaries = [' + primary + ']\n'
    f3Zone = wb.get_sheet_by_name('F3 Zone Info')
    for i in range(2,f3Zone.max_row + 1):
        zone = "".join(f3Zone.cell(row=i, column=1).value.split()).lower()
        primary = f3Zone.cell(row=i, column=2).value.strip(' ').lower()
        answer += 'let ' + zone + 'Primaries = [' + primary + ']\n'
    return answer

def createSecondaryPoints(wb) -> str:
    answer = '\n'
    d = {}
    f1ps = wb.get_sheet_by_name('F1 Primary-Secon. Edge Values')
    for i in range(2,f1ps.max_row + 1):
        primary = f1ps.cell(row=i, column=1).value
        secondary = f1ps.cell(row=i, column=2).value
        edgeValue = f1ps.cell(row=i, column=3).value
        if (secondary not in d):
            d[secondary] = []
        d[secondary].append((primary,edgeValue))
    for k,v in d.items():
        answer += 'let ' + k.lower() + 'p = ['
        for primaryEdge in v:
            answer += primaryEdge[0].lower() + ' : ' + str("{0:0.1f}".format(primaryEdge[1])) + ', '
        answer = answer[:-2] + ']\n'
    d = {}
    f2ps = wb.get_sheet_by_name('F2 Primary-Secon. Edge Values')
    for i in range(2,f2ps.max_row + 1):
        primary = f2ps.cell(row=i, column=1).value
        secondary = f2ps.cell(row=i, column=2).value
        edgeValue = f2ps.cell(row=i, column=3).value
        if (secondary not in d):
            d[secondary] = []
        d[secondary].append((primary,edgeValue))
    for k,v in d.items():
        answer += 'let ' + k.lower() + 'p = ['
        for primaryEdge in v:
            answer += primaryEdge[0].lower() + ' : ' + str("{0:0.1f}".format(primaryEdge[1])) + ', '
        answer = answer[:-2] + ']\n'
    d = {}
    f3ps = wb.get_sheet_by_name('F3 Primary-Secon. Edge Values')
    for i in range(2,f3ps.max_row + 1):
        primary = f3ps.cell(row=i, column=1).value
        secondary = f3ps.cell(row=i, column=2).value
        edgeValue = f3ps.cell(row=i, column=3).value
        if (secondary not in d):
            d[secondary] = []
        d[secondary].append((primary,edgeValue))
    for k,v in d.items():
        answer += 'let ' + k.lower() + 'p = ['
        for primaryEdge in v:
            answer += primaryEdge[0].lower() + ' : ' + str("{0:0.1f}".format(primaryEdge[1])) + ', '
        answer = answer[:-2] + ']\n'
        
    return answer

def createSecondaryExit(wb) -> str:
    answer = '\n'
    f1s = wb.get_sheet_by_name('F1 Secondary List')
    for i in range(2,f1s.max_row + 1):
        secondary = f1s.cell(row=i, column=1).value
        name = f1s.cell(row=i, column=4).value
        answer += 'let ' + secondary.lower() + ' = SecondaryExit(id: "' + secondary + '", locationName: "' + name + '", toPrimaryMap: ' + secondary.lower() + 'p)\n'
    f2s = wb.get_sheet_by_name('F2 Secondary List')
    for i in range(2,f2s.max_row + 1):
        secondary = f2s.cell(row=i, column=1).value
        name = f2s.cell(row=i, column=4).value
        answer += 'let ' + secondary.lower() + ' = SecondaryExit(id: "' + secondary + '", locationName: "' + name + '", toPrimaryMap: ' + secondary.lower() + 'p)\n'
    f3s = wb.get_sheet_by_name('F3 Secondary List')
    for i in range(2,f3s.max_row + 1):
        secondary = f3s.cell(row=i, column=1).value
        name = f3s.cell(row=i, column=4).value
        answer += 'let ' + secondary.lower() + ' = SecondaryExit(id: "' + secondary + '", locationName: "' + name + '", toPrimaryMap: ' + secondary.lower() + 'p)\n'
    return answer

def createZoneSecondaries(wb) -> str:
    answer = '\n'
    f1z = wb.get_sheet_by_name('F1 Zone Info')
    for i in range(2,f1z.max_row + 1):
        zone = "".join(f1z.cell(row=i, column=1).value.split()).lower()
        secondary = f1z.cell(row=i, column=3).value.strip(' ').lower()
        if (secondary.strip().lower() == 'n/a'):
            answer += 'let ' + zone + 'Secondaries = [SecondaryExit]()\n'
        else:
           answer += 'let ' + zone + 'Secondaries = [' + secondary + ']\n'
    f2z = wb.get_sheet_by_name('F2 Zone Info')
    for i in range(2,f2z.max_row + 1):
        zone = "".join(f2z.cell(row=i, column=1).value.split()).lower()
        secondary = f2z.cell(row=i, column=3).value.strip(' ').lower()
        if (secondary.strip().lower() == 'n/a'):
            answer += 'let ' + zone + 'Secondaries = [SecondaryExit]()\n'
        else:
           answer += 'let ' + zone + 'Secondaries = [' + secondary + ']\n'
    f3z = wb.get_sheet_by_name('F3 Zone Info')
    for i in range(2,f3z.max_row + 1):
        zone = "".join(f3z.cell(row=i, column=1).value.split()).lower()
        secondary = f3z.cell(row=i, column=3).value.strip(' ').lower()
        if (secondary.strip().lower() == 'n/a'):
            answer += 'let ' + zone + 'Secondaries = [SecondaryExit]()\n'
        else:
           answer += 'let ' + zone + 'Secondaries = [' + secondary + ']\n'
    return answer

def createZone(wb) -> str:
    answer = '\n'
    zoneList = []
    f1z = wb.get_sheet_by_name('F1 Zone Info')
    for i in range(2,f1z.max_row + 1):
        zone = f1z.cell(row=i, column=1).value
        name = "".join(zone.split()).lower()
        zoneList.append(name)
        answer += 'let ' + name + ' = Zone(name: "' + zone + '", primaryExits: ' + name + 'Primaries, secondaryExits: ' + name + 'Secondaries)\n'
    f2z = wb.get_sheet_by_name('F2 Zone Info')
    for i in range(2,f2z.max_row + 1):
        zone = f2z.cell(row=i, column=1).value
        name = "".join(zone.split()).lower()
        zoneList.append(name)
        answer += 'let ' + name + ' = Zone(name: "' + zone + '", primaryExits: ' + name + 'Primaries, secondaryExits: ' + name + 'Secondaries)\n'
    f3z = wb.get_sheet_by_name('F3 Zone Info')
    for i in range(2,f3z.max_row + 1):
        zone = f3z.cell(row=i, column=1).value
        name = "".join(zone.split()).lower()
        zoneList.append(name)
        answer += 'let ' + name + ' = Zone(name: "' + zone + '", primaryExits: ' + name + 'Primaries, secondaryExits: ' + name + 'Secondaries)\n'

    answer += '\nself.listOfZones = ['
    for z in zoneList:
        answer += z + ','
    answer = answer[:-1] + ']\n'

    return answer

if __name__ == '__main__':
    wb = openpyxl.load_workbook('Wayfinder Master Spreadsheet.xlsx')
    pe = createPrimaryExit(wb)
    zp = createZonePrimaries(wb)
    sp = createSecondaryPoints(wb)
    se = createSecondaryExit(wb)
    zs = createZoneSecondaries(wb)
    z = createZone(wb)
    file = open('IndoorSearchSimulator.txt','w')
    file.write(pe)
    file.write(zp)
    file.write(sp)
    file.write(se)
    file.write(zs)
    file.write(z)
    file.close()
