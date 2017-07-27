import openpyxl

def getPrimaries(wb) -> str:
    answer = "var graph: WeightedGraph<String, Double> = WeightedGraph<String, Double>(vertices: ["
    primaries = set()
    f1Sheet = wb.get_sheet_by_name('F1 Primary List')
    for i in range(2,f1Sheet.max_row+1):
        if (f1Sheet.cell(row=i, column=1).value != None):
            for p in f1Sheet.cell(row=i, column=1).value.split(','):
                primaries.add(p)
    f2Sheet = wb.get_sheet_by_name('F2 Primary List')
    for i in range(2,f2Sheet.max_row+1):
        if (f2Sheet.cell(row=i, column=1).value != None):
            for p in f2Sheet.cell(row=i, column=1).value.split(','):
                primaries.add(p)
    f3Sheet = wb.get_sheet_by_name('F3 Primary List')
    for i in range(2,f3Sheet.max_row+1):
        if (f3Sheet.cell(row=i, column=1).value != None):
            for p in f3Sheet.cell(row=i, column=1).value.split(','):
                primaries.add(p)
    primaries = sorted(list(primaries))
    for p in primaries:
        answer += '"' + p + '",'
    answer = answer[:-1] + '])\n\n'
    return answer

def f1PrimaryPrimaryDistance(wb) -> str:
    answer = ""
    ppSheet = wb.get_sheet_by_name('F1 Primary-Primary Edge Values')
    for i in range(2,ppSheet.max_row+1):
        if (ppSheet.cell(row=i, column=1).value != None):
            answer += 'graph.addEdge(from: "' + ppSheet.cell(row=i, column=1).value + '", to: "' + ppSheet.cell(row=i, column=2).value + '", weight: ' + str("{0:0.1f}".format(ppSheet.cell(row=i, column=3).value)) + ')\n'
    return answer

def f2PrimaryPrimaryDistance(wb) -> str:
    answer = ""
    ppSheet = wb.get_sheet_by_name('F2 Primary-Primary Edge Values')
    for i in range(2,ppSheet.max_row+1):
        if (ppSheet.cell(row=i, column=1).value != None):
            answer += 'graph.addEdge(from: "' + ppSheet.cell(row=i, column=1).value + '", to: "' + ppSheet.cell(row=i, column=2).value + '", weight: ' + str("{0:0.1f}".format(ppSheet.cell(row=i, column=3).value)) + ')\n'
    return answer

def f3PrimaryPrimaryDistance(wb) -> str:
    answer = ""
    ppSheet = wb.get_sheet_by_name('F3 Primary-Primary Edge Values')
    for i in range(2,ppSheet.max_row+1):
        if (ppSheet.cell(row=i, column=1).value != None):
            answer += 'graph.addEdge(from: "' + ppSheet.cell(row=i, column=1).value + '", to: "' + ppSheet.cell(row=i, column=2).value + '", weight: ' + str("{0:0.1f}".format(ppSheet.cell(row=i, column=3).value)) + ')\n'
    return answer

if __name__ == '__main__':
    wb = openpyxl.load_workbook('Wayfinder Master Spreadsheet.xlsx')
    p = getPrimaries(wb)
    f1 = f1PrimaryPrimaryDistance(wb)
    f2 = f2PrimaryPrimaryDistance(wb)
    f3 = f3PrimaryPrimaryDistance(wb)
    file = open('Searcher.txt','w')
    file.write(p)
    file.write(f1)
    file.write(f2)
    file.write(f3)
    file.close()
